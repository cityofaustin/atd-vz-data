from openpyxl import load_workbook
import re
import json
import os
import psycopg2
import psycopg2.extras
import onepasswordconnectsdk
from sshtunnel import SSHTunnelForwarder
from onepasswordconnectsdk.client import Client, new_client
import tempfile
import shutil

DEPLOYMENT_ENVIRONMENT = os.environ.get(
    "ENVIRONMENT", "development"
)  # our current environment from ['production', 'development']
ONEPASSWORD_CONNECT_TOKEN = os.getenv("OP_API_TOKEN")  # our secret to get secrets 🤐
ONEPASSWORD_CONNECT_HOST = os.getenv("OP_CONNECT")  # where we get our secrets
VAULT_ID = os.getenv("OP_VAULT_ID")

def main():
    global DB_HOST
    global DB_USER
    global DB_PASS
    global DB_NAME
    global DB_SSL_REQUIREMENT

    global DB_BASTION_HOST_SSH_USERNAME
    global DB_BASTION_HOST_SSH_PRIVATE_KEY
    global DB_BASTION_HOST
    global DB_RDS_HOST

    secrets = get_secrets()

    DB_HOST = secrets["database_host"]
    DB_USER = secrets["database_username"]
    DB_PASS = secrets["database_password"]
    DB_NAME = secrets["database_name"]
    DB_SSL_REQUIREMENT = secrets["database_ssl_policy"]

    DB_BASTION_HOST_SSH_USERNAME = secrets["bastion_ssh_username"]
    DB_BASTION_HOST_SSH_PRIVATE_KEY = secrets["bastion_ssh_private_key"]
    DB_BASTION_HOST = secrets["bastion_host"]
    DB_RDS_HOST = secrets["database_host"]

    print("DB_BASTION_HOST: ", DB_BASTION_HOST)

    # read_xlsx_to_get_FK_relationships("/data/cris_spec.xlsx")
    create_cris_lookup_tables("/data/cris_spec.xlsx")


def create_cris_lookup_tables(file_path):
    with SshKeyTempDir() as key_directory:
        write_key_to_file(key_directory + "/id_ed25519", DB_BASTION_HOST_SSH_PRIVATE_KEY + "\n") 
        ssh_tunnel = SSHTunnelForwarder(
            (DB_BASTION_HOST),
            ssh_username=DB_BASTION_HOST_SSH_USERNAME,
            ssh_private_key=f"{key_directory}/id_ed25519",
            remote_bind_address=(DB_RDS_HOST, 5432),
        )
        ssh_tunnel.start()

        pg = psycopg2.connect(
            host="localhost",
            port=ssh_tunnel.local_bind_port,
            user=DB_USER,
            password=DB_PASS,
            dbname=DB_NAME,
            sslmode=DB_SSL_REQUIREMENT,
            sslrootcert="/root/rds-combined-ca-bundle.pem",
        )

        workbook = load_workbook(filename=file_path)
        for worksheet in workbook.worksheets:
            print("")
            print("Title: ", worksheet.title.lower())
            match = re.search(r"(\w+)_LKP", worksheet.title)
            lookup_table = match.group(1).lower() if match else None
            if lookup_table:
                if not lookup_table == 'veh_mod_year':
                    pass
                    #continue

                skip_tables = ("cntl_sect", "inv_notify_meth")
                #skip_tables = ("cas_transp_name", "cas_transp_locat", "ins_co_name", "cntl_sect", "inv_notify_meth") # for dev
                if lookup_table in skip_tables:
                    continue

                drop = f"drop table if exists cris_lookup.{lookup_table} cascade;"
                drop_cursor = pg.cursor()
                print(f"Drop: {drop}")
                drop_cursor.execute(drop)
                drop_cursor.close()
                pg.commit()

                if lookup_table == "state":
                    populate_state_table(worksheet, lookup_table, pg)
                elif lookup_table == "veh_mod_year":
                    populate_veh_mod_year_table(worksheet, lookup_table, pg)
                else:
                    populate_table(worksheet, lookup_table, pg)

def populate_table(worksheet, lookup_table, pg):
    print("Lookup Table: ", lookup_table)

    create = f"""create table cris_lookup.{lookup_table} (
        id serial primary key, 
        upstream_id integer, 
        description text, 
        effective_begin_date date, 
        effective_end_date date,
        active boolean default true
        );"""

    create_cursor = pg.cursor()
    print(f"Create: {create}")
    create_cursor.execute(create)
    create_cursor.close()
    pg.commit()

    for row in worksheet.iter_rows(values_only=True, min_row=2):
        print(row)
        insert = f"""insert into cris_lookup.{lookup_table}
            (upstream_id, description, effective_begin_date, effective_end_date)
            values (%s, %s, %s, %s);"""
        insert_cursor = pg.cursor()
        print(f"Insert: {insert}")
        insert_cursor.execute(insert, row[:4])
        insert_cursor.close()
        pg.commit()


def populate_state_table(worksheet, lookup_table, pg):
    print("Lookup Table: ", lookup_table)

    create = f"""create table cris_lookup.{lookup_table} (
        id serial primary key, 
        upstream_id integer, 
        abbreviation text,
        description text, 
        effective_begin_date date, 
        effective_end_date date,
        active boolean default true
        );"""

    create_cursor = pg.cursor()
    print(f"Create: {create}")
    create_cursor.execute(create)
    create_cursor.close()
    pg.commit()

    for row in worksheet.iter_rows(values_only=True, min_row=2):
        print(row)
        insert = f"""insert into cris_lookup.{lookup_table}
            (upstream_id, abbreviation, description, effective_begin_date, effective_end_date)
            values (%s, %s, %s, %s, %s);"""
        insert_cursor = pg.cursor()
        print(f"Insert: {insert}")
        insert_cursor.execute(insert, row[:5])
        insert_cursor.close()
        pg.commit()

def populate_veh_mod_year_table(worksheet, lookup_table, pg):
    print("Lookup Table: ", lookup_table)


    create = f"""create table cris_lookup.{lookup_table} (
        id serial primary key, 
        upstream_id integer, 
        description text, 
        active boolean default true
        );"""

    create_cursor = pg.cursor()
    print(f"Create: {create}")
    create_cursor.execute(create)
    create_cursor.close()
    pg.commit()

    for row in worksheet.iter_rows(values_only=True, min_row=2):
        print(row)
        insert = f"""insert into cris_lookup.{lookup_table}
            (upstream_id, description)
            values (%s, %s);"""
        insert_cursor = pg.cursor()
        print(f"Insert: {insert}")
        insert_cursor.execute(insert, row[:2])
        insert_cursor.close()
        pg.commit()


def process_worksheet(worksheet, lookups):
    """
    Process a worksheet based on title specification and update the lookups dictionary.
    """
    # print("Title: ", worksheet.title.lower())

    for row in worksheet.iter_rows(values_only=True, min_row=9):
        if "lookup" in str(row[10]).lower():
            # print("")
            match = re.search(r"#'(\w+)_LKP'", row[9])
            lookup_table = match.group(1).lower() if match else None
            field = str(row[7]).split(".")[1].lower()

            # print(f"Lookup Table: '{lookup_table}'")
            # print(f"Field: '{field}'")
            lookups[field] = lookup_table


def read_xlsx_to_get_FK_relationships(file_path):
    # Load the workbook
    workbook = load_workbook(filename=file_path)

    crash_lookups = dict()
    unit_lookups = dict()
    person_lookups = dict()
    primaryperson_lookups = dict()

    for worksheet in workbook.worksheets:
        if worksheet.title.lower() == "crash file specification":
            process_worksheet(worksheet, crash_lookups)
        if worksheet.title.lower() == "unit file specification":
            process_worksheet(worksheet, unit_lookups)
        if worksheet.title.lower() == "person file specification":
            process_worksheet(worksheet, person_lookups)
        if worksheet.title.lower() == "primaryperson file spec.":
            process_worksheet(worksheet, primaryperson_lookups)

    print(json.dumps(crash_lookups, indent=4))
    print(json.dumps(unit_lookups, indent=4))
    print(json.dumps(person_lookups, indent=4))
    print(json.dumps(primaryperson_lookups, indent=4))


def get_secrets():
    REQUIRED_SECRETS = {
        "bastion_host": {
            "opitem": "RDS Bastion Host",
            "opfield": f"{DEPLOYMENT_ENVIRONMENT}.Host",
            "opvault": VAULT_ID,
        },
        "bastion_ssh_username": {
            "opitem": "RDS Bastion Host",
            "opfield": f"{DEPLOYMENT_ENVIRONMENT}.ssh Username",
            "opvault": VAULT_ID,
        },
        "database_host": {
            "opitem": "Vision Zero Database",
            "opfield": f"{DEPLOYMENT_ENVIRONMENT}.Database Host",
            "opvault": VAULT_ID,
        },
        "database_username": {
            "opitem": "Vision Zero Database",
            "opfield": f"{DEPLOYMENT_ENVIRONMENT}.Database Username",
            "opvault": VAULT_ID,
        },
        "database_password": {
            "opitem": "Vision Zero Database",
            "opfield": f"{DEPLOYMENT_ENVIRONMENT}.Database Password",
            "opvault": VAULT_ID,
        },
        "database_name": {
            "opitem": "Vision Zero Database",
            "opfield": f"{DEPLOYMENT_ENVIRONMENT}.Database Name",
            "opvault": VAULT_ID,
        },
        "database_ssl_policy": {
            "opitem": "Vision Zero Database",
            "opfield": f"{DEPLOYMENT_ENVIRONMENT}.Database SSL Policy",
            "opvault": VAULT_ID,
        },
        "bastion_ssh_private_key": {
            "opitem": "RDS Bastion Key",
            "opfield": ".private key",
            "opvault": VAULT_ID,
            },
    }

    # instantiate a 1Password client
    client: Client = new_client(ONEPASSWORD_CONNECT_HOST, ONEPASSWORD_CONNECT_TOKEN)
    # get the requested secrets from 1Password
    return onepasswordconnectsdk.load_dict(client, REQUIRED_SECRETS)


# these temp directories are used to store ssh keys, because they will
# automatically clean themselves up when they go out of scope.
class SshKeyTempDir:
    def __init__(self):
        self.path = None

    def __enter__(self):
        self.path = tempfile.mkdtemp(dir='/tmp')
        return self.path

    def __exit__(self, exc_type, exc_val, exc_tb):
        shutil.rmtree(self.path)

def write_key_to_file(path, content):
    # Open the file with write permissions and create it if it doesn't exist
    fd = os.open(path, os.O_WRONLY | os.O_CREAT, 0o600)

    # Write the content to the file
    os.write(fd, content.encode())

    # Close the file
    os.close(fd)



if __name__ == "__main__":
    main()